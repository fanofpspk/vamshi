import win32com.client as client
import openpyxl as op
import re
from bs4 import BeautifulSoup
import time
from datetime import datetime, timedelta

# Use 'with open' to read the configuration file
config_file = "config.txt"  # Path to your configuration file
with open(config_file, "r") as file:
    content = file.read()
    # Use 'exec' to load the variables from the content of the file
    exec(content)

# Now you can access these variables directly:
print("Loaded configuration:")
print(f"HR: {HR}")
print(f"Feedback senders: {Feedback_senders}")
print(f"Job IDs: {job_id}")
print(f"Status Times: {Status_times}")

# Define the Excel file path
excel_file = "candidate_status1.xlsx"  # Path of your Excel file

# Try loading the existing workbook, or create a new one if it doesn't exist
try:
    wb = op.load_workbook(excel_file)  # Try loading the existing workbook
    print(f"Workbook '{excel_file}' loaded successfully.")
except FileNotFoundError:
    wb = op.Workbook()  # If the file doesn't exist, create a new one
    print(f"Workbook '{excel_file}' not found. Creating a new one.")

# Ensure the sheet exists, or create it
if "candidate_status2" not in wb.sheetnames:
    ws1 = wb.create_sheet("candidate_status2")
    ws1.append(["Name", "mblNumber", "Experience", "CTC", "EXPECTED_CTC", "CURRENT_CTC", "NOTICE_PERIOD", "Status","Job_ID"])
    print("New sheet 'candidate_status2' created.")
else:
    ws1 = wb["candidate_status2"]

# New Sheet to Track Forwarded Candidates:
# Create or load the sheet to track forwarded candidates
if "forwarded_candidates" not in wb.sheetnames:
    forwarded_sheet = wb.create_sheet("forwarded_candidates")
    forwarded_sheet.append(["Name"])  # Only store candidate names
    print("New sheet 'forwarded_candidates' created.")
else:
    forwarded_sheet = wb["forwarded_candidates"]

# Initialize Outlook application and MAPI namespace
outlook = client.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.Folders.Item("vamshichandra5677@outlook.com").Folders.Item("Inbox")

# Store candidate details temporarily before forwarding and receiving reply
candidate_details = {}

# Fetch already forwarded candidates from the 'forwarded_candidates' sheet
forwarded_candidates = set()
for row in forwarded_sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:
        forwarded_candidates.add(row[0].lower())  # Store in lowercase for case-insensitive comparison

# Status map for status updates
status_map = {
    "selected": 1,
    "rejected": 5,
    "pending": 3,
}
jd_missing_sent = set()

# To track when the forward email was sent for a candidate
forward_time = {}  # Track the forward time of each candidate

# To track when reminder was sent for a candidate
reminder_time = {}

# Set to track candidates who have replied and had status updated
replied_candidates = set()

# To track the number of reminders sent for each candidate
reminder_count = {}

# To track the last reminder sent time for each candidate
last_reminder_time = {}

# Time period between 1:30 AM to 1:35 AM
pause_start_time = timedelta(hours=4, minutes=25)
pause_end_time = timedelta(hours=4, minutes=30)

# Now use a delay to check new emails every X seconds (e.g., 60 seconds)
while True:
    # Fetch all messages in the inbox (Get only the latest one)
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)  # Sort by most recent emails

    # Process the latest email
    for msg in messages:
        if hasattr(msg, 'SenderEmailAddress') and msg.SenderEmailAddress.lower() in HR:
            if not any(job_id in msg.Subject for job_id in job_id):
                print(f"Email with mismatched subject '{msg.Subject}' found. Deleting it.")
                msg.Delete()  # Delete email if subject doesn't match valid ones
                continue

            # Convert body text to lowercase for consistent comparison
            body = msg.Body.strip().lower()

            # Check if "jd" or "Job description" is in the email body
            if "jd" not in body and "job description" not in body:
                print("No 'jd' or 'Job description' found in email body. Responding to sender.")

                # Extract the candidate's name from the email (for reference)
                candidate_name = None
                html_body = msg.HTMLBody
                soup = BeautifulSoup(html_body, "html.parser")
                table = soup.find("table")
                if table:
                    for row in table.find_all("tr")[1:]:
                        cols = row.find_all("td")
                        if len(cols) == 7:
                            candidate_name = cols[0].get_text(strip=True)
                            break

                # Send "JD missing" reply only once per candidate
                if candidate_name and candidate_name.lower() not in jd_missing_sent:
                    # Reply to the sender
                    reply_mail = msg.Reply()
                    reply_mail.Subject = "Re: " + msg.Subject
                    reply_mail.Body = "No JD (Job Description) found in your email. Please provide more information."
                    reply_mail.Send()
                    msg.Delete()
                    continue

            else:
                print("JD detected in email body")

                # Extract the HTML body of the email
                html_body = msg.HTMLBody
                soup = BeautifulSoup(html_body, "html.parser")

                # Find the table in the email
                table = soup.find("table")
                if table:
                    # Iterate over each row in the table
                    for row in table.find_all("tr")[1:]:
                        cols = row.find_all("td")
                        if len(cols) == 7:
                            data = [col.get_text(strip=True) for col in cols]

                            # Store candidate details for later use
                            candidate_name = data[0]
                            candidate_details[candidate_name] = {
                                "Name": data[0],
                                "mblNumber": data[1],
                                "Experience": data[2],
                                "CTC": data[3],
                                "EXPECTED_CTC": data[4],
                                "CURRENT_CTC": data[5],
                                "NOTICE_PERIOD": data[6],
                                "Status": 3,
                                "Job ID": msg.Subject.split()[0]# Default to pending
                            }

                            # Check if candidate already exists in the Excel sheet
                            candidate_found = False
                            for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, values_only=False):
                                if row[0].value and row[0].value.lower() == candidate_name.lower():
                                    candidate_found = True
                                    next_row = row[0].row  # Get the row of the existing candidate
                                    break

                            if candidate_found:
                                ws1.cell(row=next_row, column=8, value=3)  # Update status
                                print(f"Status updated for existing candidate '{candidate_name}'.")
                            else:
                                next_row = ws1.max_row + 1
                                for idx, value in enumerate(data):
                                    ws1.cell(row=next_row, column=idx + 1, value=value)
                                ws1.cell(row=next_row, column=8, value=3)  # Status column
                                ws1.cell(row=next_row, column=9,
                                         value=candidate_details[candidate_name]["Job ID"])  # Job ID column
                                forward_time[candidate_name] = datetime.now()
                                print(f"Data appended for new candidate: {', '.join(data)}")

                else:
                    print("No table found in email body.")

    # Save the workbook with updated data
    wb.save(excel_file)
    print(f"Data saved to '{excel_file}' successfully.")

    # Check for replies and send reminders if needed
    current_time = datetime.now()

    # Monitor replies to forwarded emails and update status
    for msg in messages:
        if hasattr(msg, 'SenderEmailAddress'):
            if msg.SenderEmailAddress.lower() == "vadderajeshvarma0406@gmail.com".lower() and msg.Subject.startswith(
                    "Re:"):
                print(f"Reply Subject: {msg.Subject}")
                print(f"Reply From: {msg.SenderName}")
                print(f"Reply Received: {msg.ReceivedTime}")
                print(f"Reply Body: {msg.Body}")
                print("=" * 50)

                # Convert the reply body to lowercase for easier keyword matching
                reply_body = msg.Body.lower()

                # Try to find the candidate name in the reply
                candidate_name = None
                for name in candidate_details:
                    if re.search(r'\b' + re.escape(name.lower()) + r'\b', reply_body):  # Match whole words only
                        candidate_name = name
                        break

                # If the candidate's name is found, update the status
                if candidate_name:
                    print(f"Candidate '{candidate_name}' found in reply body. Updating status.")

                    # Determine the status from the reply body
                    status = 3  # Default to pending
                    for keyword, status_value in status_map.items():
                        if keyword in reply_body:
                            status = status_value
                            break

                    # Find the candidate's row in the Excel sheet and update the status
                    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, values_only=False):
                        if row[0].value.lower() == candidate_name.lower():
                            ws1.cell(row=row[0].row, column=8, value=status)  # Update status
                            print(f"Status updated for '{candidate_name}' to {status} at row {row[0].row}.")
                            break

                    # Mark the candidate as replied and no further reminder should be sent
                    replied_candidates.add(candidate_name)
                    print(f"Candidate '{candidate_name}' has replied. Status updated, no further reminder will be sent.")

                    # Stop sending reminder if the status is updated
                    reminder_time[candidate_name] = current_time  # Mark the reminder time as updated
                    print(f"Reminder time updated for '{candidate_name}'. No further reminders will be sent.")

    # Reminder logic: Send reminders if candidates haven't replied
    for candidate_name in candidate_details:
        if candidate_name not in replied_candidates:
            # Initialize reminder count if not present
            if candidate_name not in reminder_count:
                reminder_count[candidate_name] = 0

            # Check for pause time (1:30 AM to 1:35 AM)
            pause_time = current_time.time()
            pause_start = datetime.combine(datetime.today(), datetime.min.time()) + pause_start_time
            pause_end = datetime.combine(datetime.today(), datetime.min.time()) + pause_end_time

            # Send reminder only if it's not during the pause period
            if not (pause_start.time() <= pause_time <= pause_end.time()):
                # Send reminder at correct intervals (every 2 minutes)
                reminder_interval = timedelta(minutes=2)  # Fixed interval of 2 minutes

                # Only send a reminder if the correct interval has passed since the last reminder
                if candidate_name in forward_time and current_time - forward_time[
                        candidate_name] > reminder_interval * (reminder_count.get(candidate_name, 0) + 1):
                    #reminder_mail = client.Dispatch("Outlook.Application").CreateItem(0)
                    #outlook =  client.Dispatch("Outlook.Application").GetNamespace("MAPI")
                    print(outlook)  # Check if Outlook application is initialized
                    outlook = client.Dispatch("Outlook.Application")
                    namespace = outlook.GetNamespace("MAPI")
                    reminder_mail = outlook.CreateItem(0)  # Create a new mail item

                    reminder_mail.To = "vadderajeshvarma0406@gmail.com"
                    reminder_mail.Subject = f"Reminder {reminder_count[candidate_name] + 1}: Status for candidate {candidate_name}"
                    reminder_mail.Body = f"Please provide the status for candidate {candidate_name}. This is reminder {reminder_count[candidate_name] + 1}."
                    reminder_mail.Send()

                    # Update the reminder time and increment the reminder count
                    reminder_time[candidate_name] = current_time
                    reminder_count[candidate_name] += 1
                    print(f"Reminder {reminder_count[candidate_name]} sent for candidate {candidate_name}.")

    # 9:00 PM email: Send a daily status summary email and create a new Excel file
    current_time = datetime.now()
    if current_time.hour == 10 and current_time.minute == 59:  # 9:00 PM
        daily_status_html = """
                <h2>Daily Embedded Requirements Sourcing Report</h2>
                <table border='1'>
                    <tr>
                        <th>Job Code</th>
                        <th>Sourced</th>
                        <th>Shortlisted</th>
                        <th>Rejected</th>
                        <th>Pending</th>
                        <th>Screen Remarks</th>
                        <th>Sourcing Efficiency (%)</th>
                    </tr>
                """

        # Initialize job_status dynamically from Excel data
        job_status = {}  # Format: {job_id: [sourced, shortlisted, rejected, pending]}
        remarks = {}  # Format: {job_id: [remarks]}

        # Process Excel data
        for row in ws1.iter_rows(min_row=2, values_only=True):
            if len(row) >= 9 and row[8]:  # Validate Job ID exists
                job_id = row[8]
                status = row[7]

                # Initialize job_id in dictionaries if not present
                if job_id not in job_status:
                    job_status[job_id] = [0, 0, 0, 0]  # sourced, shortlisted, rejected, pending
                    remarks[job_id] = []

                # Update counts based on status
                job_status[job_id][0] += 1  # Sourced
                if status == 1:  # Shortlisted
                    job_status[job_id][1] += 1
                elif status == 5:  # Rejected
                    job_status[job_id][2] += 1
                    remarks[job_id].append(f"{row[0]} - Rejected")
                else:  # Pending
                    job_status[job_id][3] += 1

        # Build HTML table
        for job_id, counts in job_status.items():
            sourced, shortlisted, rejected, pending = counts
            efficiency = round((shortlisted / sourced * 100), 2) if sourced > 0 else 0
            screen_remarks = "<br>".join(remarks[job_id]) or "No remarks"

            daily_status_html += f"""
                    <tr>
                        <td>{job_id}</td>
                        <td>{sourced}</td>
                        <td>{shortlisted}</td>
                        <td>{rejected}</td>
                        <td>{pending}</td>
                        <td>{screen_remarks}</td>
                        <td>{efficiency}%</td>
                    </tr>
                    """

        daily_status_html += "</table>"

        # Send email
        mail = outlook.Application.CreateItem(0)
        mail.Subject = "Daily Embedded Requirements Sourcing Report"
        mail.HTMLBody = daily_status_html
        mail.To = "vadderaj06@gmail.com"  # Corrected email
        mail.Send()
        print("Email sent with data.")

    for row in ws1.iter_rows(values_only=True):
        print(row)
    # print('Sleeping for 10 seconds')  # Add a short pause to prevent unnecessary high CPU usage
    time.sleep(10)  # Check every minute